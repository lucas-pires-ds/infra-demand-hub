"""
generate_all.py
Gera dados fictícios para o projeto infra-demand-hub.
Fontes: API (qualyteam), CSV (Planner), EML (E-mail), XLSX (Call)
"""

import csv
import json
import random
import os
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(7)

DOMINIO = "portocentral.com.br"
EMAIL_INFRA = f"infra@{DOMINIO}"

NOW = datetime(2025, 5, 2, 10, 0, 0)

RESPONSAVEIS = [
    "Carlos Menezes",
    "Fernanda Lopes",
    "Roberto Aguiar",
    "Patrícia Souza",
    "Diego Fonseca",
    "Leandro Braga",
]

RESPONSAVEIS_ABREV = {
    "Carlos Menezes":  "Carlos M.",
    "Fernanda Lopes":  "Fernanda L.",
    "Roberto Aguiar":  "Roberto A.",
    "Patrícia Souza":  "Patrícia S.",
    "Diego Fonseca":   "Diego F.",
    "Leandro Braga":   "Leandro B.",
}

DEMANDAS_POOL = [
    # (titulo, categoria, descricao, prioridade, sla_horas)

    ("Inspeção estrutural Cais 2 – defensas e vigas",                "Estrutura Civil",           "Cais 2 apresenta deformação visível em defensa lateral após atracação de navio de grande porte. Inspeção técnica necessária antes de liberar novas atracações.",                                                                              "Crítica", 8),
    ("Verificação fundação – base guindaste MHC-07",                 "Engenharia / Laudo Técnico","Base de fixação do guindaste MHC-07 com indício de recalque diferencial. Operação do equipamento suspensa até laudo de engenharia civil.",                                                                                                  "Crítica", 12),
    ("Falha em quadro elétrico QF-03 – Galpão 1",                    "Elétrica",                  "Quadro de força QF-03 do Galpão 1 apresentou desarmamento de disjuntor geral. Setor sem energia. Equipe de manutenção elétrica solicitada com urgência.",                                                                                   "Crítica", 4),
    ("Iluminação pátio noturno setor C com falha",                   "Elétrica / Iluminação",     "Setor C do pátio de operações com 4 postes de iluminação inoperantes. Operação noturna comprometida e risco de acidente. Turno noturno em andamento.",                                                                                      "Crítica", 3),
    ("Cabo de aço guindaste MHC-04 com desgaste crítico",            "Manutenção Equipamentos",   "Inspeção identificou fios rompidos no cabo de içamento do guindaste MHC-04. Desgaste acima do limite normativo. Operação suspensa por segurança.",                                                                                          "Crítica", 6),
    ("Chiller do servidor de controle com alarme de temperatura",    "Climatização / HVAC",       "Sistema de refrigeração do rack de controle operacional com alarme ativo. Temperatura interna do ambiente subindo. Risco para equipamentos.",                                                                                               "Crítica", 2),
    ("Câmera 14 Cais Principal offline",                             "CFTV / Segurança",          "Câmera 14 do Cais Principal offline desde as 13h20. Ponto cego na área de maior movimentação do terminal. Turno tarde e noite sem cobertura no local.",                                                                                     "Crítica", 4),
    ("Cancela eletrônica Portaria 2 travada em posição fechada",     "Controle de Acesso",        "Cancela eletrônica da Portaria 2 travou em posição fechada impossibilitando entrada de veículos. Fila de caminhões se formando. Portaria operando em modo manual.",                                                                         "Crítica", 2),
    ("Sinalização de emergência Galpão 3 inoperante",                "Sinalização / Segurança",   "Sistema de sinalização de emergência do Galpão 3 não está funcionando. Saídas de emergência sem indicação luminosa. Não conformidade grave com normas de segurança.",                                                                       "Crítica", 6),
    ("Portão principal travado – acesso bloqueado",                  "Controle de Acesso",        "Portão de acesso principal travou em posição fechada. Equipes externas e fornecedores sem acesso ao terminal. Operação de descarga programada para 14h em risco.",                                                                          "Crítica", 2),

    ("Laudo capacidade de solo – Pátio Norte expansão RTG",          "Engenharia / Laudo Técnico","Solicitação de laudo técnico de capacidade de carga do solo na área de expansão do Pátio Norte para instalação de guindaste RTG de 320 toneladas. Necessário antes da chegada do equipamento prevista para 15/05.",                        "Alta",    360),
    ("Reforma cobertura Galpão 2 – infiltração generalizada",        "Estrutura Civil",           "Cobertura do Galpão 2 com múltiplos pontos de infiltração comprometendo carga armazenada. Necessário projeto e execução de reforma completa.",                                                                                              "Alta",    240),
    ("Adequação piso Pátio Sul para tráfego de reach stacker",       "Estrutura Civil",           "Piso do Pátio Sul com trincas e afundamentos em área de manobra de reach stacker. Necessário estudo de reforço estrutural e recapeamento.",                                                                                                 "Alta",    120),
    ("Gerador G2 não partiu no teste semanal",                       "Elétrica / Gerador",        "Gerador G2 (500 kVA) não deu partida durante teste semanal programado. Alarme de falha ativo no painel. Gerador G1 operando normalmente. Risco para continuidade em caso de falta de energia.",                                            "Alta",    24),
    ("SPDA com haste rompida – Galpão 4",                            "Elétrica / SPDA",           "Sistema de proteção contra descargas atmosféricas do Galpão 4 com haste coletora rompida após tempestade. Galpão sem proteção adequada.",                                                                                                  "Alta",    48),
    ("Empilhadeira EMP-09 sem freio de estacionamento",              "Manutenção Equipamentos",   "Operador reportou falha no freio de estacionamento da empilhadeira EMP-09. Equipamento imobilizado por risco de acidente.",                                                                                                                 "Alta",    12),
    ("Correia transportadora TRS-02 com desgaste",                   "Manutenção Equipamentos",   "Correia transportadora TRS-02 do terminal de granéis sólidos com desgaste acentuado nas bordas. Substituição necessária antes da próxima operação de descarga.",                                                                            "Alta",    24),
    ("Vazamento na rede hidráulica – Galpão 3",                      "Hidráulica",                "Vazamento visível na tubulação de água fria no corredor lateral do Galpão 3. Água acumulando no piso e risco de danos a equipamentos.",                                                                                                     "Alta",    8),
    ("Sistema de drenagem Pátio Sul entupido",                       "Hidráulica / Drenagem",     "Canaletas de drenagem do Pátio Sul completamente obstruídas. Com previsão de chuvas para o fim de semana, risco de alagamento da área operacional.",                                                                                        "Alta",    24),
    ("Reparo calçada Portaria Principal – risco de queda",           "Estrutura Civil",           "Calçada na entrada da Portaria Principal com piso solto e desnível. Colaborador sofreu quase-acidente. Área sinalizada, reparo urgente necessário.",                                                                                        "Alta",    12),
    ("Ar condicionado sala de operações inoperante",                 "Climatização / HVAC",       "Split da sala de operações (térreo) parou de funcionar. Sala com temperatura acima de 35°C. Operadores em condições inadequadas de trabalho. Equipamento tem 5 anos de uso.",                                                              "Alta",    8),
    ("Cerca perimetral com ruptura – setor fundo do pátio",          "Segurança Física",          "Inspeção de ronda identificou ruptura de aproximadamente 3 metros na cerca perimetral no fundo do Pátio Norte. Área sensível próxima ao armazém alfandegado.",                                                                             "Alta",    12),
    ("Portão de acesso veicular Galpão 1 com defeito",               "Controle de Acesso",        "Portão de correr do Galpão 1 saiu dos trilhos e não abre completamente. Acesso de caminhões para carga e descarga prejudicado.",                                                                                                           "Alta",    8),
    ("Bebedouro refeitório com vazamento interno",                   "Facilities / Hidráulica",   "Bebedouro industrial do refeitório com vazamento interno. Piso molhado e risco de acidente. Bebedouro desligado até reparo.",                                                                                                              "Alta",    8),

    ("Impermeabilização área de armazenagem Galpão 5",               "Estrutura Civil",           "Galpão 5 com laje apresentando infiltração após chuvas fortes. Risco de danos a equipamentos armazenados e comprometimento estrutural.",                                                                                                    "Média",   96),
    ("Adequação elétrica para novo compressor de ar industrial",     "Elétrica",                  "Necessidade de adequação do ramal elétrico na área de manutenção para suportar novo compressor de ar industrial (380V / 50A). Projeto e execução.",                                                                                         "Média",   72),
    ("Manutenção preventiva – 4 empilhadeiras programada",           "Manutenção Preventiva",     "Manutenção preventiva trimestral das empilhadeiras EMP-01, EMP-03, EMP-05 e EMP-08. Troca de óleo, filtros, verificação de freios e sistema hidráulico.",                                                                                  "Média",   48),
    ("Pintura de sinalização pátio desgastada",                      "Sinalização / Segurança",   "Faixas de sinalização de circulação de veículos e pedestres do pátio operacional completamente desgastadas. Não conformidade identificada em auditoria de segurança.",                                                                      "Média",   120),
    ("Ar condicionado sala de reuniões 2° andar com ruído",          "Climatização / HVAC",       "Ar condicionado da sala de reuniões do 2° andar fazendo barulho e sem resfriamento eficiente desde segunda-feira. Segunda ocorrência no mês.",                                                                                             "Média",   48),
    ("Banheiro masculino térreo – vaso sanitário entupido",          "Facilities / Hidráulica",   "Vaso sanitário do banheiro masculino do térreo entupido e transbordando. Banheiro interditado. Colaboradores utilizando apenas o do 2° andar.",                                                                                            "Média",   4),
    ("Manutenção preventiva CFTV – limpeza e foco câmeras",          "CFTV / Manutenção Preventiva","Manutenção preventiva semestral do sistema de CFTV: limpeza de lentes, ajuste de foco, verificação de gravação e teste de todas as 34 câmeras.",                                                                                        "Média",   72),
    ("Engenharia / Laudo Técnico – sondagem Pátio Leste",            "Engenharia / Laudo Técnico","Solicitação de sondagem SPT na área do Pátio Leste para verificação de capacidade de suporte antes de ampliação da área de armazenagem de contêineres.",                                                                                   "Média",   72),

    ("Lâmpadas queimadas corredor administrativo 1° andar",          "Elétrica / Iluminação",     "Três luminárias do corredor do administrativo 1° andar com lâmpadas queimadas. Ambiente com iluminação insuficiente.",                                                                                                                     "Baixa",   48),
    ("Reparo de calha – Galpão 4 lateral direita",                   "Estrutura Civil",           "Calha de escoamento de água pluvial do Galpão 4 com trecho solto e entupido. Água transbordando para a lateral do galpão em dias de chuva.",                                                                                               "Baixa",   96),
    ("Torneira do vestiário masculino com vazamento",                 "Facilities / Hidráulica",   "Torneira do vestiário masculino com vazamento contínuo. Desperdício de água.",                                                                                                                                                             "Baixa",   48),
    ("Revisão de fechaduras – salas administrativas",                "Facilities / Hidráulica",   "Três fechaduras de salas administrativas com dificuldade de abertura. Necessário lubrificação e ajuste.",                                                                                                                                  "Baixa",   72),
]

SOLICITANTES = [
    ("Rogério Andrade",       "Operações Portuárias",  f"rogerio.andrade@{DOMINIO}"),
    ("Marcos Vinicius Teles", "Planejamento",           f"mv.teles@{DOMINIO}"),
    ("Jorge Salles",          "Segurança Patrimonial",  f"jorge.salles@{DOMINIO}"),
    ("Paula Drummond",        "Administrativo",         f"paula.drummond@{DOMINIO}"),
    ("Antônio Ferreira",      "Manutenção",             f"antonio.ferreira@{DOMINIO}"),
    ("Claudia Nascimento",    "Administrativo",         f"claudia.nascimento@{DOMINIO}"),
    ("Sérgio Pimentel",       "Operações Portuárias",   f"sergio.pimentel@{DOMINIO}"),
    ("Beatriz Carvalho",      "RH",                     f"beatriz.carvalho@{DOMINIO}"),
    ("Fábio Gomes",           "Manutenção",             f"fabio.gomes@{DOMINIO}"),
    ("Renata Moura",          "Financeiro",             f"renata.moura@{DOMINIO}"),
    ("Thiago Lima",           "Almoxarifado",           f"thiago.lima@{DOMINIO}"),
    ("Cristiane Barros",      "Compliance",             f"cristiane.barros@{DOMINIO}"),
]


def random_date_past(days_ago_max=10, days_ago_min=0):
    delta = random.randint(days_ago_min * 3600, days_ago_max * 24 * 3600)
    return NOW - timedelta(seconds=delta)


def prazo_from_abertura(abertura, sla_horas):
    return abertura + timedelta(hours=sla_horas)


def data_conclusao(status, abertura, prazo):
    if status == "Concluído":
        if random.random() < 0.75:
            delta = (prazo - abertura).total_seconds()
            return abertura + timedelta(seconds=random.randint(int(delta * 0.3), int(delta * 0.95)))
        else:
            return prazo + timedelta(hours=random.randint(1, 12))
    return None


def assign_status(prioridade, prazo):
    vencido = prazo < NOW
    if prioridade == "Crítica":
        if vencido:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.35, 0.25, 0.40])[0]
        else:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.30, 0.45, 0.25])[0]
    elif prioridade == "Alta":
        if vencido:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.25, 0.20, 0.55])[0]
        else:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.25, 0.45, 0.30])[0]
    else:
        if vencido:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.15, 0.15, 0.70])[0]
        else:
            return random.choices(["Pendente", "Em Andamento", "Concluído"], weights=[0.20, 0.40, 0.40])[0]


def distribuir_demandas(total, pool):
    criticas  = [d for d in pool if d[3] == "Crítica"]
    altas     = [d for d in pool if d[3] == "Alta"]
    medias    = [d for d in pool if d[3] == "Média"]
    baixas    = [d for d in pool if d[3] == "Baixa"]

    n_critica = round(total * 0.20)
    n_alta    = round(total * 0.35)
    n_media   = round(total * 0.30)
    n_baixa   = total - n_critica - n_alta - n_media

    def sample_rep(lst, n):
        result = []
        while len(result) < n:
            result += random.sample(lst, min(len(lst), n - len(result)))
        return result[:n]

    selecionadas = (
        sample_rep(criticas, n_critica) +
        sample_rep(altas,    n_alta)    +
        sample_rep(medias,   n_media)   +
        sample_rep(baixas,   n_baixa)
    )
    random.shuffle(selecionadas)
    return selecionadas


# ══════════════════════════════════════════════════════════════
# FONTE 1 — qualyteam (JSON via API)
# ══════════════════════════════════════════════════════════════

def gerar_qualyteam():
    registros = []
    pool = distribuir_demandas(30, DEMANDAS_POOL)

    for i, (titulo, categoria, descricao, prioridade, sla_horas) in enumerate(pool):
        solicitante = random.choice(SOLICITANTES)
        responsavel = random.choice(RESPONSAVEIS)
        abertura    = random_date_past(days_ago_max=8)
        prazo       = prazo_from_abertura(abertura, sla_horas)
        status      = assign_status(prioridade, prazo)
        conclusao   = data_conclusao(status, abertura, prazo)

        if random.random() < 0.10:
            descricao = ""
        if random.random() < 0.05:
            responsavel = None

        registros.append({
            "id_chamado":        f"QT-2025-{str(i + 280).zfill(4)}",
            "titulo":            titulo,
            "descricao":         descricao,
            "categoria":         categoria,
            "prioridade":        prioridade,
            "status":            status,
            "solicitante":       solicitante[0],
            "setor_solicitante": solicitante[1],
            "responsavel_infra": responsavel,
            "data_abertura":     abertura.strftime("%Y-%m-%dT%H:%M:%S"),
            "prazo_conclusao":   prazo.strftime("%Y-%m-%dT%H:%M:%S"),
            "data_conclusao":    conclusao.strftime("%Y-%m-%dT%H:%M:%S") if conclusao else None,
            "sla_horas":         sla_horas,
        })

    return registros


# ══════════════════════════════════════════════════════════════
# FONTE 2 — Planner/Teams (CSV)
# ══════════════════════════════════════════════════════════════

def gerar_planner_csv():
    BUCKETS = ["Manutenção Corretiva", "Manutenção Preventiva", "Obras e Reformas", "Engenharia", "", "Facilities", "Segurança"]

    rows = []
    pool = distribuir_demandas(25, DEMANDAS_POOL)

    for i, (titulo, categoria, descricao, prioridade, sla_horas) in enumerate(pool):
        responsavel = random.choice(RESPONSAVEIS)
        abertura    = random_date_past(days_ago_max=8)
        prazo       = prazo_from_abertura(abertura, sla_horas)
        status_real = assign_status(prioridade, prazo)

        if status_real == "Concluído":
            status_planner = random.choice(["Concluída", "concluída", "Concluido"])
        elif status_real == "Em Andamento":
            status_planner = random.choice(["Em andamento", "Em Andamento"])
        else:
            status_planner = random.choice(["Não iniciada", "nao iniciada"])

        pct = 100 if status_real == "Concluído" else (random.choice([25, 50, 75]) if status_real == "Em Andamento" else 0)

        if prioridade == "Crítica":
            prio_planner = random.choice(["Urgente", "urgente"])
        elif prioridade == "Alta":
            prio_planner = random.choice(["Alta", "alta"])
        elif prioridade == "Média":
            prio_planner = random.choice(["Normal", "normal"])
        else:
            prio_planner = random.choice(["Baixa", ""])

        rows.append({
            "ID_Tarefa":      f"PLAN-{str(i + 70).zfill(4)}",
            "Nome da Tarefa": titulo,
            "Bucket":         random.choice(BUCKETS) if random.random() > 0.20 else "",
            "Atribuído a":    "" if random.random() < 0.15 else responsavel,
            "Data Início":    abertura.strftime("%d/%m/%Y"),
            "Data Limite":    prazo.strftime("%d/%m/%Y"),
            "% Concluído":    pct,
            "Prioridade":     prio_planner,
            "Status":         status_planner,
            "Observações":    descricao[:60] + "..." if random.random() > 0.25 and descricao else "",
        })

    return rows


# ══════════════════════════════════════════════════════════════
# FONTE 3 — E-mail (.eml)
# ══════════════════════════════════════════════════════════════

ASSUNTOS_EMAIL = [
    ("URGENTE - Câmera 14 do Cais offline",               "Segurança Patrimonial", f"jorge.salles@{DOMINIO}",       "Jorge Salles",          "Crítica"),
    ("Ar condicionado sala de reuniões 2° andar",         "Administrativo",        f"paula.drummond@{DOMINIO}",     "Paula Drummond",        "Baixa"),
    ("Infiltração teto Galpão 3 - precisa verificar",     "Operações Portuárias",  f"sergio.pimentel@{DOMINIO}",    "Sérgio Pimentel",       "Alta"),
    ("Fwd: Portão Galpão 1 com problema - 3° aviso",      "Manutenção",            f"fabio.gomes@{DOMINIO}",        "Fábio Gomes",           "Alta"),
    ("iluminação do pátio C apagada de novo",             "Operações Portuárias",  f"rogerio.andrade@{DOMINIO}",    "Rogério Andrade",       "Crítica"),
    ("Solicitação de manutenção - bebedouro refeitório",  "RH",                    f"beatriz.carvalho@{DOMINIO}",   "Beatriz Carvalho",      "Média"),
    ("RE: Gerador G2 - sem retorno desde quinta",         "Manutenção",            f"antonio.ferreira@{DOMINIO}",   "Antônio Ferreira",      "Alta"),
    ("banheiro térreo entupido",                          "Administrativo",        f"claudia.nascimento@{DOMINIO}", "Claudia Nascimento",    "Média"),
    ("Cerca perimetral fundo do pátio com problema",      "Segurança Patrimonial", f"jorge.salles@{DOMINIO}",       "Jorge Salles",          "Alta"),
    ("Solicitação pintura sinalização pátio - auditoria", "Compliance",            f"cristiane.barros@{DOMINIO}",   "Cristiane Barros",      "Média"),
    ("URGENTE: Chiller rack controle com alarme",         "Operações Portuárias",  f"sergio.pimentel@{DOMINIO}",    "Sérgio Pimentel",       "Crítica"),
    ("lâmpadas corredor admin queimadas",                 "Administrativo",        f"renata.moura@{DOMINIO}",       "Renata Moura",          "Baixa"),
    ("Trinca na viga galpão 3 - área interditada",        "Operações Portuárias",  f"rogerio.andrade@{DOMINIO}",    "Rogério Andrade",       "Alta"),
    ("Empilhadeira EMP-09 sem freio - imobilizada",       "Manutenção",            f"fabio.gomes@{DOMINIO}",        "Fábio Gomes",           "Crítica"),
    ("Calçada portaria solta - quase acidente hoje",      "Segurança Patrimonial", f"jorge.salles@{DOMINIO}",       "Jorge Salles",          "Alta"),
    ("Solicitação laudo solo pátio norte",                "Planejamento",          f"mv.teles@{DOMINIO}",           "Marcos Vinicius Teles", "Alta"),
    ("drenagem pátio sul entupida - chuva prevista",      "Operações Portuárias",  f"antonio.ferreira@{DOMINIO}",   "Antônio Ferreira",      "Alta"),
    ("SPDA galpão 4 com haste quebrada",                  "Manutenção",            f"fabio.gomes@{DOMINIO}",        "Fábio Gomes",           "Alta"),
    ("ac sala de operações parou - calor absurdo",        "Operações Portuárias",  f"rogerio.andrade@{DOMINIO}",    "Rogério Andrade",       "Alta"),
    ("Cancela portaria 2 travada - fila de caminhões",    "Operações Portuárias",  f"sergio.pimentel@{DOMINIO}",    "Sérgio Pimentel",       "Crítica"),
]

CORPOS_EMAIL = {
    "URGENTE - Câmera 14 do Cais offline":
        "Boa tarde,\n\nA câmera 14 localizada no Cais Principal está offline desde as 13h20. A área ficará sem monitoramento durante o turno da tarde e da noite.\n\nPrecisamos de reparo o quanto antes. Operação 24h sem cobertura de CFTV é inaceitável.\n\nAtt,\nJorge Salles\nSegurança Patrimonial",
    "Ar condicionado sala de reuniões 2° andar":
        "Olá,\n\nO ar condicionado da sala de reuniões do 2° andar está fazendo barulho e não está gelando direito desde segunda-feira. Já é a segunda vez esse mês.\n\nPodem verificar? Temos uma reunião importante amanhã às 9h.\n\nObrigada,\nPaula Drummond\nAdministrativo",
    "Infiltração teto Galpão 3 - precisa verificar":
        "Pessoal,\n\nTemos infiltração no teto do Galpão 3, área central, próximo à coluna C-12. Depois das chuvas de ontem piorou bastante. Tem carga paletizada embaixo que pode ser afetada.\n\nUrgente verificar.\n\nSérgio Pimentel\nOp. Portuárias",
    "Fwd: Portão Galpão 1 com problema - 3° aviso":
        "Encaminhando novamente. Esse chamado foi aberto pela 3ª vez e ainda sem solução.\n\nO portão do Galpão 1 continua saindo dos trilhos. Ontem precisamos de 3 pessoas para abrir manualmente e atrasar a operação de descarga por 40 minutos.\n\nFábio Gomes\nManutenção",
    "iluminação do pátio C apagada de novo":
        "Chefe,\n\nO pátio C tá no escuro de novo. 4 postes apagados. A operação noturna não pode parar mas tá arriscado sem iluminação.\n\nJá reportei semana passada e voltou o problema. Aguardo providências.\n\nRogério\nOp. Portuárias",
    "Solicitação de manutenção - bebedouro refeitório":
        "Bom dia,\n\nO bebedouro industrial do refeitório está vazando. O chão ao redor está sempre molhado, já tivemos um colaborador que quase escorregou hoje cedo.\n\nDesligamos o bebedouro por segurança. Precisamos de manutenção o quanto antes pois é o único bebedouro do refeitório.\n\nBeatriz Carvalho\nRH",
    "RE: Gerador G2 - sem retorno desde quinta":
        "Preciso de um retorno sobre o gerador G2.\n\nAbri chamado na quinta-feira e até agora nenhuma resposta. O G1 está operando sozinho e qualquer falha nele nos deixa sem backup.\n\nPor favor, priorizar.\n\nAntônio Ferreira\nManutenção",
    "banheiro térreo entupido":
        "Oi,\n\nbanheiro masculino do térreo entupido e transbordando. interditei mas tá ruim pq o pessoal do térreo tem que subir tudo pro 2° andar.\n\npode resolver hoje?\n\nClaudia",
    "Cerca perimetral fundo do pátio com problema":
        "Equipe Infra,\n\nDurante a ronda das 06h identificamos ruptura na cerca perimetral no fundo do Pátio Norte. São aproximadamente 3 metros de cerca caída. A área é próxima ao armazém alfandegado, o que é grave do ponto de vista de segurança e compliance.\n\nNecessário reparo urgente e registro formal da ocorrência.\n\nJorge Salles\nSegurança Patrimonial",
    "Solicitação pintura sinalização pátio - auditoria":
        "Equipe,\n\nDurante a auditoria interna de segurança realizada ontem foi identificado que as faixas de sinalização do pátio operacional estão completamente desgastadas e em alguns pontos invisíveis.\n\nIsso configura não conformidade com a NR-12 e precisa ser corrigido antes da auditoria externa prevista para o mês que vem.\n\nCristiane Barros\nCompliance",
    "URGENTE: Chiller rack controle com alarme":
        "ATENÇÃO,\n\nO sistema de refrigeração do rack de controle operacional disparou alarme de temperatura alta. A temperatura interna está subindo e risco real para os equipamentos de controle do terminal.\n\nNecessário intervenção IMEDIATA.\n\nSérgio Pimentel\nOp. Portuárias",
    "lâmpadas corredor admin queimadas":
        "Boa tarde,\n\nTrês luminárias do corredor do administrativo no 1° andar com lâmpadas queimadas. Corredor escuro. Se puder resolver essa semana fica ótimo.\n\nRenata\nFinanceiro",
    "Trinca na viga galpão 3 - área interditada":
        "Urgente,\n\nOperador de empilhadeira identificou trinca visível na viga principal do Galpão 3 próximo à coluna C-7. Interditar a área preventivamente até inspeção estrutural.\n\nRogério Andrade",
    "Empilhadeira EMP-09 sem freio - imobilizada":
        "Bom dia,\n\nA empilhadeira EMP-09 foi imobilizada pelo operador após falha no freio de estacionamento. Equipamento parado na área de carga do Galpão 2.\n\nNecessário reparo urgente pois estamos com frota reduzida.\n\nFábio Gomes",
    "Calçada portaria solta - quase acidente hoje":
        "Equipe,\n\nQuase acidente grave hoje pela manhã na Portaria Principal. Um colaborador tropeçou no piso solto da calçada de entrada. Área sinalizada com cone mas não é suficiente.\n\nPrecisamos resolver hoje ainda.\n\nJorge Salles",
    "Solicitação laudo solo pátio norte":
        "Equipe Infra,\n\nPrecisamos formalizar a solicitação de laudo técnico de capacidade de carga do solo na área de expansão do Pátio Norte.\n\nEstamos prevendo a instalação de um guindaste RTG de 320 toneladas e o fornecedor exige o laudo antes da entrega, prevista para 15/05. Prazo apertado.\n\nMarcos Vinicius Teles\nPlanejamento",
    "drenagem pátio sul entupida - chuva prevista":
        "Pessoal,\n\nAs canaletas de drenagem do Pátio Sul estão completamente obstruídas. A previsão é de chuvas fortes no fim de semana e com essa situação o alagamento é certo.\n\nPrecisamos desentupir antes de sexta.\n\nAntônio Ferreira",
    "SPDA galpão 4 com haste quebrada":
        "Infra,\n\nApós a tempestade de ontem identifiquei que a haste coletora do SPDA do Galpão 4 está quebrada. O galpão está sem proteção adequada contra raios.\n\nFábio Gomes",
    "ac sala de operações parou - calor absurdo":
        "Equipe,\n\nO split da sala de operações do térreo parou completamente. Está passando de 35 graus aqui dentro. Os operadores não estão conseguindo trabalhar direito.\n\nEquipamento tem uns 5 anos, talvez precise de substituição.\n\nRogério",
    "Cancela portaria 2 travada - fila de caminhões":
        "URGENTE\n\nCancela eletrônica da Portaria 2 travou fechada. Tem fila de caminhões se formando na rua. Estamos operando em modo manual mas é insustentável.\n\nNecessário técnico agora.\n\nSérgio Pimentel",
}


def gerar_emls(output_dir):
    os.makedirs(output_dir, exist_ok=True)

    for i, (assunto, setor, email_from, nome, prioridade) in enumerate(ASSUNTOS_EMAIL):
        abertura = random_date_past(days_ago_max=8)

        msg = MIMEMultipart()
        msg["From"]       = f"{nome} <{email_from}>"
        msg["To"]         = EMAIL_INFRA
        msg["Subject"]    = assunto
        msg["Date"]       = abertura.strftime("%a, %d %b %Y %H:%M:%S -0300")
        msg["Message-ID"] = f"<PC{str(i+1000).zfill(6)}.{abertura.strftime('%Y%m%d')}@{DOMINIO}>"

        corpo = CORPOS_EMAIL.get(assunto, f"Solicitação de manutenção: {assunto}.\n\nAguardo retorno.\n\n{nome}\n{setor}")
        msg.attach(MIMEText(corpo, "plain", "utf-8"))

        filename = f"email_{str(i+1).zfill(3)}_{assunto[:40].replace(' ', '_').replace('/', '-').replace(':', '')}.eml"
        with open(os.path.join(output_dir, filename), "w", encoding="utf-8") as f:
            f.write(msg.as_string())

    print(f"  ✓ {len(ASSUNTOS_EMAIL)} arquivos .eml gerados em {output_dir}")


# ══════════════════════════════════════════════════════════════
# FONTE 4 — Call/Telefone (XLSX)
# ══════════════════════════════════════════════════════════════

CHAMADOS_CALL = [
    ("Antônio Ferreira",   "Manutenção",           "9988-4421",  "Gerador G2 não partiu no teste semanal. Alarme ativo no painel.",                 "URGENTE", "Carlos M.",   "Em andamento"),
    ("Claudia Nascimento", "Administrativo",        "97654-3310", "Infiltração no teto da sala de arquivos, próximo à janela.",                      "media",   "Patrícia S.", "concluido"),
    ("Sérgio Pimentel",    "Operações Portuárias",  "3421-9900",  "Pátio C sem iluminação. 4 postes apagados. Operação noturna parada.",             "URGENTE", "Roberto A.",  "Em andamento"),
    ("Jorge Salles",       "Segurança",             "98877-0012", "Câmera do cais offline. Precisa de técnico hoje.",                                 "urgente", "Roberto A.",  "Concluído"),
    ("Fábio Gomes",        "Manutenção",            "3421-9901",  "Empilhadeira EMP-09 com defeito no freio. Imobilizada.",                          "URGENTE", "Fernanda L.", "Em aberto"),
    ("Paula Drummond",     "Administrativo",        "99123-4455", "AC da sala de reuniões com barulho.",                                             "baixa",   "Patrícia S.", "Concluído"),
    ("Rogério Andrade",    "Operações Portuárias",  "98234-5566", "Portão do Galpão 1 fora dos trilhos de novo.",                                    "Alta",    "Leandro B.",  "Em aberto"),
    ("Beatriz Carvalho",   "RH",                    "3421-0023",  "Bebedouro do refeitório vazando. Piso molhado.",                                  "Media",   "Patrícia S.", "concluido"),
    ("Marcos Teles",       "Planejamento",          "99887-2211", "Laudo solo para o RTG novo. Urgente, prazo do fornecedor.",                       "Alta",    "Diego F.",    "Em andamento"),
    ("Antônio Ferreira",   "Manutenção",            "9988-4421",  "Cabo de aço do MHC-04 com fios rompidos. Operação suspensa.",                    "URGENTE", "Carlos M.",   "Concluído"),
    ("Cristiane Barros",   "Compliance",            "3421-9905",  "Sinalização do pátio apagada. Auditoria apontou não conformidade.",               "normal",  "",            "Pendente"),
    ("Thiago Lima",        "Almoxarifado",          "97001-3344", "Portão de acesso do almoxarifado com dobradiça quebrada.",                        "media",   "Leandro B.",  "concluido"),
    ("Sérgio Pimentel",    "Operações Portuárias",  "3421-9900",  "Chiller do rack de controle disparou alarme de temperatura.",                     "URGENTE", "Roberto A.",  "Em andamento"),
    ("Jorge Salles",       "Segurança",             "98877-0012", "Cerca perimetral fundo do pátio rasgada. Aprox 3 metros.",                        "Alta",    "Diego F.",    "Em andamento"),
    ("Paula Drummond",     "Administrativo",        "99123-4455", "Banheiro masculino entupido e transbordando.",                                    "media",   "Patrícia S.", "Concluído"),
    ("Rogério Andrade",    "Operações Portuárias",  "98234-5566", "AC da sala de operações parou. Calor insuportável.",                              "URGENTE", "Fernanda L.", "Em aberto"),
    ("Antônio Ferreira",   "Manutenção",            "9988-4421",  "Drenagem pátio sul entupida. Chuva prevista sexta.",                             "Alta",    "Carlos M.",   "Em aberto"),
    ("Fábio Gomes",        "Manutenção",            "3421-9901",  "SPDA galpão 4 com haste quebrada após tempestade.",                              "Alta",    "Diego F.",    "Pendente"),
    ("Claudia Nascimento", "Administrativo",        "97654-3310", "Lâmpadas do corredor admin queimadas. Corredor escuro.",                          "Baixa",   "",            "Pendente"),
    ("Sérgio Pimentel",    "Operações Portuárias",  "3421-9900",  "Cancela portaria 2 travada fechada. Fila de caminhões na rua.",                   "URGENTE", "Roberto A.",  "Em andamento"),
    ("Thiago Lima",        "Almoxarifado",          "97001-3344", "Vazamento na rede de água corredor lateral galpão 3.",                           "Alta",    "Leandro B.",  "Em andamento"),
    ("Beatriz Carvalho",   "RH",                    "3421-0023",  "Trinca na viga do galpão 3. Área interditada.",                                  "URGENTE", "Carlos M.",   "Em aberto"),
]

HORAS_CALL = ["07:30", "08:00", "08:15", "09:00", "09:45", "10:20", "11:00",
              "13:30", "14:00", "14:30", "15:15", "16:00", "16:45", "17:00",
              "18:30", "19:00", "20:15", "22:00", "23:30", "00:15", "05:45", "06:00"]


def gerar_xlsx_call(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro de Chamados"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ["Nº", "Data", "Hora", "Quem Ligou", "Setor", "Descrição do Problema",
               "Urgência", "Responsável", "Status", "Tel Contato", "Obs"]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [5, 12, 8, 22, 22, 55, 10, 16, 14, 14, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    base_date = NOW - timedelta(days=7)
    current_date = base_date

    for i, (quem, setor, tel, descricao, urgencia, responsavel, status) in enumerate(CHAMADOS_CALL):
        current_date += timedelta(hours=random.randint(2, 12))
        if current_date > NOW:
            current_date = NOW - timedelta(hours=random.randint(1, 3))

        hora = random.choice(HORAS_CALL)

        obs = ""
        if status in ["concluido", "Concluído"] and random.random() > 0.3:
            obs = f"resolvido {current_date.strftime('%d/%m')} às {random.choice(HORAS_CALL)}"
        elif random.random() < 0.15:
            obs = "cliente ligou novamente cobrando"

        ws.append([
            i + 1,
            current_date.strftime("%d/%m/%Y"),
            hora,
            quem, setor, descricao, urgencia, responsavel, status, tel, obs,
        ])

        if i % 2 == 0:
            fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
            for cell in ws[i + 2]:
                cell.fill = fill

    wb.save(filepath)
    print(f"  ✓ XLSX salvo em {filepath} ({len(CHAMADOS_CALL)} registros)")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    base   = os.path.dirname(os.path.abspath(__file__))
    output = os.path.join(base, "output")
    os.makedirs(output, exist_ok=True)

    print("\n🏗️  Gerando dados fictícios — infra-demand-hub\n")

    print("📡 [1/4] qualyteam (JSON para API)...")
    qt_data = gerar_qualyteam()
    qt_path = os.path.join(output, "qualyteam_demandas.json")
    with open(qt_path, "w", encoding="utf-8") as f:
        json.dump(qt_data, f, ensure_ascii=False, indent=2)
    print(f"  ✓ JSON salvo em {qt_path} ({len(qt_data)} registros)")

    print("📋 [2/4] Planner/Teams (CSV)...")
    rows = gerar_planner_csv()
    csv_path = os.path.join(output, "planner_tarefas.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✓ CSV salvo em {csv_path} ({len(rows)} registros)")

    print("📧 [3/4] E-mails (.eml)...")
    gerar_emls(os.path.join(output, "emails_eml"))

    print("📞 [4/4] Call/Telefone (XLSX)...")
    gerar_xlsx_call(os.path.join(output, "call_chamados.xlsx"))

    print("\n✅ Todos os dados gerados com sucesso!")
    print(f"   Pasta: {output}\n")
